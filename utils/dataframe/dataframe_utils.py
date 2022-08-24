import os.path
import re
import pandas as pd
import numpy as np
import json
import logging
from collections import OrderedDict
from utils.regex_utils import check_compile_regex
from utils.debug_utils import print_file_function
from utils.text.lines import get_multiline_post_para_offsets, get_matches_with_group_relative_offsets,\
    combine_matches_with_post_groups, print_combined_matches, print_matches_with_post_groups, \
    extend_match_groups_with_post_groups, set_groups_absolute_offset
from utils.regex_utils import regex_apply_on_text


logger = logging.getLogger(__name__)


def is_dataframe(var):
    return isinstance(var, pd.DataFrame)


def df_new_dataframe():
    return pd.DataFrame()


def create_df_from_text_using_regex(regex_text, input_file_text, flags=None):
    p, error = check_compile_regex(regex_text, flags=flags)

    if error:
        logger.error("Regex has errors: ", error)
        return None

    s = pd.Series(input_file_text)
    try:
        df = s.str.extractall(regex_text , re.MULTILINE)
    except ValueError as e:
        logger.error(e)

    # Verify for later:
    # Drop level 0 as we are using the whole string and the level was create as we converted string to pd.Series
    df = df.droplevel(0)

    return df


def create_dataframe_from_combined_matches(matches):
    records = []
    for m_idx,m in enumerate(matches):
        print("match[{}]".format(m_idx))
        rec = OrderedDict()
        for g_idx,g in enumerate(m['groups']):
            print("group[{}:{}]:\n{}\n{}".format(g_idx, g['name'], g['text'], g['offsets_list']))
            rec[g['name']] = g['text']
        records.append(rec)

    df = pd.DataFrame(records)
    return df


def create_dataframe_from_text_extrapolate(regex_str, input_str, flags=None, extrapolate=False, extp_join_str="\n"):
    result = regex_apply_on_text(regex_str, input_str, flags=flags)

    if extrapolate:
        matches = result['matches']

        multiline_matches = get_multiline_post_para_offsets(matches, len(input_str))

        print("Matches with post para")
        for m in multiline_matches:
            print(m)

        matches_with_post_groups = get_matches_with_group_relative_offsets(input_str, multiline_matches)

        # print_matches_with_post_groups(matches_with_post_groups)

        # The extended groups are good for display for they loose the information required for combining
        matches_with_extended_groups = extend_match_groups_with_post_groups(matches_with_post_groups)

        print("Matches with Extended Groups:")
        for m in multiline_matches:
            print(m)

        matches_with_absolute_offsets = set_groups_absolute_offset(matches_with_extended_groups)
        print("Matches with Absolute Offset in Groups:")
        for m in multiline_matches:
            print(m)

        matches_combined = combine_matches_with_post_groups(matches_with_post_groups, join_str=extp_join_str)

        # print_combined_matches(matches_combined)

        df = create_dataframe_from_combined_matches(matches_combined)
        print(type(df))
        print(df)

    return df


def df_apply_regex_on_column(df, regex_text, column=None):
    if column is None:
        raise RuntimeError("column cannot be None")

    new_df = df[column].str.extract(regex_text, expand=True)
    df[new_df.columns] = new_df

    return df


# TBD: This is a major problem here.
# We need to replace columns if all are nan
# Even though we have created this, it is not good enough
def df_merge_on_index(df1, df2, columns=[]):
    if len(columns) < 1:
        columns = df2.columns

    logger.info("df_merge_on_index(): columns={}".format(columns))

    left_suffix = "_x"
    right_suffix = "_y"
    df_merged = df1.merge(df2, how="left", left_index=True, right_index=True, suffixes=(left_suffix, right_suffix))

    left_columns = [col+left_suffix for col in columns]
    right_columns = [col+right_suffix for col in columns]

    if (set(left_columns).issubset(set(df_merged.columns))):
        logger.info("We need to merge now")

    df_print(df_merged, index=True, shape=True)

    return df_merged


def df_apply_regexlist_on_column(df, regex_list, column=None, new_anchor_column=None, remove=True, multiple=False, join_original=True, join_columns=[], debug=False):
    if column is None:
        raise RuntimeError("column cannot be None")

    if debug:
        logger.info("df_apply_regexlist_on_column(): Input frame")
        df_print(df[column])

    match_df = None
    for index, regex_text in enumerate(regex_list):
        if debug:
            logger.info("regex={}".format(regex_text))

        if not column in df.columns:
            raise RuntimeError("column '{}' not found in columns: {}".format(column, df.columns))

        if multiple:
            new_df = df[column].str.extractall(regex_text, re.MULTILINE)
            index_level_count = len(new_df.index.levels)
            if debug:
                logger.info("New DF: Index Level Count: {}".format(index_level_count))
            new_df = new_df.droplevel(index_level_count - 1)
        else:
            new_df = df[column].str.extract(regex_text, re.MULTILINE, expand=True)

        if match_df is None:
            match_df = new_df
        else:
            # Set only the rows which are not already set
            if new_anchor_column is not None:
                match_df.loc[match_df[new_anchor_column].isna(), new_df.columns] = new_df

        if debug:
            logger.info("Match DF:")
            df_print(match_df, index=True, shape=True, dtypes=True)

    if join_original:
        # Check: Force dropping of columns
        # df.drop(column, axis=1, inplace=True)

        df = df[join_columns].join(match_df, how="left", lsuffix="_x", rsuffix="_y")

        if remove and new_anchor_column is not None:
            df.loc[match_df[new_anchor_column].notna(), column] = np.nan

        if debug:
            logger.info("DF after joining:")
            df_print(df, index=True, shape=True, dtypes=True)
    else:
        df = match_df

    return df


FLAG_ACTIVE_DEFAULT = True
FLAG_FORCE_LOCATION = False

def df_print(df, count=20, dtypes=False, index=False, shape=False, new_line=True, gui=False, active=True, location=True, columns=False):
    if not active:
        return

    # https://stackoverflow.com/questions/6810999/how-to-determine-file-function-and-line-number

    # Use this as a decorator
    if location or FLAG_FORCE_LOCATION:
        print_file_function(offset=1, levels=4)

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.set_option('display.max_colwidth', None)
    pd.set_option('display.float_format', lambda x: '%.2f' % x)

    if gui:
        # gui = show(df, settings={'block': True})
        logger.info("pandas_gui not used")
    else:
        if count > 0:
            df = df.head(count)

        if new_line:
            logger.info("\n")

        logger.info(df)

        if index:
            logger.info(df.index)

        if shape:
            logger.info(df.shape)

        if dtypes:
            logger.info(df.dtypes)

        if columns:
            logger.info(df.columns)


def df_read_excel(*args, **kwargs):
    return pd.read_excel(*args, **kwargs)


def df_read_csv(*args, **kwargs):
    return pd.read_csv(*args, **kwargs)


# https://xlsxwriter.readthedocs.io/example_pandas_datetime.html
def df_write_excel(df, filepath, *args, **kwargs):
    if 'index' not in kwargs:
        kwargs['index'] = False

    writer = pd.ExcelWriter(filepath,
                            engine='xlsxwriter',
                            datetime_format='yyyy-mm-dd',
                            date_format='yyyy-mm-dd')

    df.to_excel(writer, *args, **kwargs)
    writer.save()


def dflist_write_excel(dflist, filepath, *args, **kwargs):
    if 'index' not in kwargs:
        kwargs['index'] = False

    writer = pd.ExcelWriter(filepath,
                            engine='xlsxwriter',
                            datetime_format='yyyy-mm-dd',
                            date_format='yyyy-mm-dd')

    for dfentry in dflist:
        df = dfentry['dataframe']
        suffix = dfentry['suffix']

        if suffix is None or suffix == "":
            suffix = "Sheet1"
            logger.info('dflist_write_excel: no sheetname provided, using default {}'.format(suffix))

        df.to_excel(writer, sheet_name=suffix, *args, **kwargs)

    writer.save()


def df_to_oriented_json(df):
    return json.loads(df.to_json(orient="table"))


def df_from_oriented_json(json_data):
    return pd.read_json(json.dumps(json_data), orient='table')


# https://stackoverflow.com/questions/49519696/getting-attributeerror-workbook-object-has-no-attribute-add-worksheet-whil
def df_append_excel(df, file_path, sheet_name, **kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in kwargs:
        kwargs.pop('engine')
    startrow = kwargs.pop('startrow', None)
    truncate_sheet = kwargs.pop('truncate_sheet', None)

    if 'index' not in kwargs:
        kwargs['index'] = False

    file_exists = os.path.exists(file_path)
    if file_exists:
        book = load_workbook(file_path)

    # This will create a file of size zero
    writer = pd.ExcelWriter(file_path,
                            engine='openpyxl',
                            datetime_format='yyyy-mm-dd',
                            date_format='yyyy-mm-dd')

    if file_exists:
        writer.book = book
        # logger.info("Existing sheetnames={}".format(writer.book.sheetnames))

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **kwargs)

    # save the workbook
    writer.save()

# https://stackoverflow.com/questions/20625582/how-to-deal-with-settingwithcopywarning-in-pandas
# Usage:
# with SupressSettingWithCopyWarning():
#     code that produces warning
#
class SupressSettingWithCopyWarning:
    def __enter__(self):
        pd.options.mode.chained_assignment = None

    def __exit__(self, *args):
        pd.options.mode.chained_assignment = 'warn'


type_group_map = {
    "nan": "float",
    "int": "number",
    "float": "number",
    "str": "string",
    "datetime[ns64]": "date",
    "TimeStamp": "date"
}


def get_type_group(type_str):
    return type_group_map[type_str]


# The match_strategy currently is disabled
def filter_by_signature_and_value(row, row_filter, match_strategy="exact"):
    # TBD: We might support a percentage match
    match = True
    for (cell, cell_filter) in  zip(row, row_filter):
        cell_type = get_cell_type_signature(cell)

        accepted_types = cell_filter.get("types")
        if accepted_types is None:
            accepted_type = cell_filter.get("type")
            if isinstance(accepted_type, list):
                raise RuntimeError("Only one type accepted. For multiple use types")
            accepted_types = [accepted_type]
        else:
            if not isinstance(accepted_types, list):
                raise RuntimeError("Specify types in a list format")

        if cell_type not in accepted_types:
            match = False
            break

    return match


# TBD: This seems insufficient.
# Even though we have knowledge whether filtered row is a header or not but we are not passing it back
def filter_by_row_and_header_signature_and_value(row, row_signature, header_signature=None, match_strategy="exact"):
    match = filter_by_signature_and_value(row, row_signature, match_strategy=match_strategy)

    if not match and header_signature is not None:
        match = filter_by_signature_and_value(row, header_signature, match_strategy=match_strategy)

    return match


def get_cell_type_signature(cell):
    return type(cell).__name__


def df_type_signature(df):
    return df.applymap(lambda x: get_cell_type_signature(x))


def df_filter_by_row_and_header_signature(df, row_signature, header_signature=None, match_strategy='exact', debug=False):
    logger.info("df_filter_by_row_signature():")
    logger.info("signature: {}".format(row_signature))
    logger.info("header_signature: {}".format(header_signature))

    # This is currently not used
    signature_df = df_type_signature(df)
    if debug:
        logger.info("signature_df: ")
        df_print(signature_df)

    boolean_frame = df.apply(
        lambda row: filter_by_row_and_header_signature_and_value(row,
                                                                 row_signature,
                                                                 header_signature=header_signature,
                                                                 match_strategy=match_strategy),
        axis=1
    )

    return df[boolean_frame]



def df_is_empty(df):
    return df.empty


aggregate_function_map = {
    "sum": np.sum,
}


def get_aggregate_function(key):
    return aggregate_function_map.get(key, np.sum)
